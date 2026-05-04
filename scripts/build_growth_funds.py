"""
資産運用業協会のExcelから成長投資枠対象ファンドを growth_funds.js に変換する。

使い方:
  python3 scripts/build_growth_funds.py
"""

import json
import unicodedata
import openpyxl
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent


def normalize(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", str(s).strip())
    return s


def load_growth_funds():
    wb = openpyxl.load_workbook(ROOT / "data" / "growth_unlisted.xlsx")

    # 削除済みコードを収集
    ws_del = wb["削除の予定・履歴"]
    deleted_codes = set()
    for row in list(ws_del.iter_rows(values_only=True))[1:]:
        if row[2]:
            deleted_codes.add(str(row[2]).strip())

    # 有効ファンドを収集
    ws = wb["対象商品一覧"]
    funds = []
    seen_names = set()
    for row in list(ws.iter_rows(values_only=True))[2:]:
        code = str(row[2]).strip() if row[2] else ""
        name = normalize(row[3])
        company = normalize(row[4])
        if not name or code in deleted_codes:
            continue
        if name in seen_names:
            continue
        seen_names.add(name)
        funds.append({
            "name": name,
            "company": company,
            "fundCode": code,
            "frame": "growth",
        })

    return funds


def main():
    funds = load_growth_funds()
    print(f"成長投資枠ファンド数: {len(funds)}")

    js = f"""// 成長投資枠対象商品リスト（非上場ファンド）
// 出典: 資産運用業協会「NISA成長投資枠の対象商品」
// URL: https://www.imaj.or.jp/find/nisa_growth_productslist/
// 取得日: {__import__('datetime').date.today().isoformat()}
// ライセンス: 公共データ（各社届出に基づく協会公表情報）

const growthFundDatabase = {json.dumps(funds, ensure_ascii=False, indent=2)};
"""

    out = ROOT / "data" / "growth_funds.js"
    out.write_text(js, encoding="utf-8")
    print(f"保存: {out}")
    print(f"サンプル:")
    for f in funds[:3]:
        print(f"  {f['name'][:40]} / {f['company'][:25]}")


if __name__ == "__main__":
    main()
